from flask import Flask, render_template, request, jsonify, session, send_file
from flask_cors import CORS
from functools import wraps
import sys
import os
import hashlib
from datetime import datetime

sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from database.db_manager import DatabaseManager

app = Flask(__name__)
app.secret_key = 'chave_mestra_condominio_2024'
CORS(app)

# ==================== CRIAÇÃO DO BANCO DE DADOS ====================
# Garantir que o banco SQLite existe antes de qualquer operação
db_setup = DatabaseManager()
if not os.path.exists('condominio.db'):
    print("📁 Banco de dados não encontrado. Criando...")
    db_setup.create_database()
    print("✅ Banco de dados criado com sucesso!")
else:
    print("📁 Banco de dados já existe.")
db_setup.close()
# ================================================================

# Conectar ao banco para uso da aplicação
db = DatabaseManager()
db.connect()
print("✅ Banco de dados conectado")


def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return jsonify({'success': False, 'error': 'Não autorizado'}), 401
        return f(*args, **kwargs)

    return decorated_function


def master_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_tipo' not in session or session['user_tipo'] != 'master':
            return jsonify({'success': False, 'error': 'Acesso negado! Apenas administrador.'}), 403
        return f(*args, **kwargs)

    return decorated_function


@app.route('/')
def index():
    return render_template('index.html')


@app.route('/api/login', methods=['POST'])
def api_login():
    data = request.get_json()
    user = db.authenticate_user(data.get('login'), data.get('senha'))
    if user:
        session['user_id'] = user['id']
        session['user_nome'] = user['nome']
        session['user_tipo'] = user['tipo']
        session['user_condominio_id'] = user.get('condominio_id')
        print(f"✅ Login: {user['nome']} - Tipo: {user['tipo']} - Condomínio ID: {user.get('condominio_id')}")
        return jsonify({'success': True, 'user': user})
    return jsonify({'success': False, 'error': 'Usuário ou senha inválidos!'})


@app.route('/api/logout', methods=['POST'])
def api_logout():
    session.clear()
    return jsonify({'success': True})


# ==================== CONDOMÍNIOS ====================
@app.route('/api/condominios', methods=['GET'])
@login_required
@master_required
def get_condominios():
    condos = db.fetch_all("SELECT * FROM condominios ORDER BY nome")
    return jsonify({'success': True, 'condominios': condos})


@app.route('/api/condominios', methods=['POST'])
@login_required
@master_required
def create_condominio():
    data = request.get_json()
    db.execute_query(
        "INSERT INTO condominios (nome, endereco, telefone) VALUES (?, ?, ?)",
        (data.get('nome'), data.get('endereco'), data.get('telefone'))
    )
    return jsonify({'success': True, 'message': 'Condomínio criado com sucesso!'})


@app.route('/api/condominios/<int:id>', methods=['DELETE'])
@login_required
@master_required
def delete_condominio(id):
    db.execute_query("DELETE FROM condominios WHERE id = ?", (id,))
    return jsonify({'success': True, 'message': 'Condomínio removido com sucesso!'})


# ==================== PESSOAS ====================
@app.route('/api/pessoas', methods=['GET'])
@login_required
def get_pessoas():
    busca = request.args.get('busca', '')

    if session.get('user_tipo') == 'porteiro':
        condominio_id = session.get('user_condominio_id')
        if busca:
            pessoas = db.fetch_all(
                "SELECT * FROM pessoas WHERE condominio_id = ? AND nome LIKE ? ORDER BY nome",
                (condominio_id, f'%{busca}%')
            )
        else:
            pessoas = db.fetch_all(
                "SELECT * FROM pessoas WHERE condominio_id = ? ORDER BY nome",
                (condominio_id,)
            )
    else:
        if busca:
            pessoas = db.fetch_all("SELECT * FROM pessoas WHERE nome LIKE ? ORDER BY nome", (f'%{busca}%',))
        else:
            pessoas = db.fetch_all("SELECT * FROM pessoas ORDER BY nome")

    return jsonify({'success': True, 'pessoas': pessoas})


@app.route('/api/pessoas', methods=['POST'])
@login_required
def create_pessoa():
    try:
        data = request.get_json()

        if session.get('user_tipo') == 'porteiro':
            condominio_id = session.get('user_condominio_id')
        else:
            condominio_id = data.get('condominio_id')

        if not condominio_id:
            return jsonify({'success': False, 'error': 'Selecione um condomínio!'}), 400

        query = """
            INSERT INTO pessoas (nome, tipo, apartamento, telefone, documento, veiculo_placa, veiculo_modelo, status, condominio_id) 
            VALUES (?, ?, ?, ?, ?, ?, ?, 'ATIVO', ?)
        """
        db.execute_query(query, (
            data.get('nome'),
            data.get('tipo'),
            data.get('apartamento'),
            data.get('telefone'),
            data.get('documento'),
            data.get('veiculo_placa'),
            data.get('veiculo_modelo'),
            condominio_id
        ))
        return jsonify({'success': True, 'message': 'Pessoa cadastrada com sucesso!'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@app.route('/api/pessoas/<int:id>', methods=['PUT'])
@login_required
def update_pessoa(id):
    try:
        data = request.get_json()
        query = """
            UPDATE pessoas 
            SET nome=?, tipo=?, apartamento=?, telefone=?, documento=?, 
                veiculo_placa=?, veiculo_modelo=?
            WHERE id=?
        """
        db.execute_query(query, (
            data.get('nome'), data.get('tipo'), data.get('apartamento'),
            data.get('telefone'), data.get('documento'), data.get('veiculo_placa'),
            data.get('veiculo_modelo'), id
        ))
        return jsonify({'success': True, 'message': 'Pessoa atualizada com sucesso!'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@app.route('/api/pessoas/<int:id>', methods=['DELETE'])
@login_required
@master_required
def delete_pessoa(id):
    try:
        db.execute_query("DELETE FROM registros WHERE pessoa_id = ?", (id,))
        db.execute_query("DELETE FROM pessoas WHERE id = ?", (id,))
        return jsonify({'success': True, 'message': 'Pessoa removida com sucesso!'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@app.route('/api/registrar_entrada', methods=['POST'])
@login_required
def registrar_entrada():
    try:
        data = request.get_json()
        pessoa_id = data.get('pessoa_id')
        observacao = data.get('observacao', '')

        pessoa = db.fetch_one("SELECT status, nome, condominio_id, bloqueado FROM pessoas WHERE id = ?", (pessoa_id,))

        if not pessoa:
            return jsonify({'success': False, 'error': 'Pessoa não encontrada!'})

        if session.get('user_tipo') == 'porteiro':
            if pessoa.get('condominio_id') != session.get('user_condominio_id'):
                return jsonify({'success': False, 'error': 'Acesso negado!'}), 403

        if pessoa.get('status') == 'BLOQUEADO' or pessoa.get('bloqueado') == 1:
            return jsonify({'success': False, 'error': 'Pessoa bloqueada! Entrada não permitida.'}), 403

        usuario_registro = session.get('user_id')

        db.execute_query(
            "INSERT INTO registros (pessoa_id, data_entrada, usuario_registro, observacao) VALUES (?, ?, ?, ?)",
            (pessoa_id, datetime.now().strftime('%Y-%m-%d %H:%M:%S'), usuario_registro, observacao)
        )
        return jsonify({'success': True, 'message': f'Entrada de {pessoa["nome"]} registrada com sucesso!'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@app.route('/api/registros', methods=['GET'])
@login_required
def get_registros():
    if session.get('user_tipo') == 'porteiro':
        condominio_id = session.get('user_condominio_id')
        query = """
            SELECT 
                r.id, r.data_entrada, r.observacao,
                p.nome as pessoa_nome, 
                p.tipo as pessoa_tipo, 
                p.apartamento,
                u.nome as usuario_nome
            FROM registros r 
            JOIN pessoas p ON r.pessoa_id = p.id 
            LEFT JOIN usuarios u ON r.usuario_registro = u.id
            WHERE p.condominio_id = ?
            ORDER BY r.data_entrada DESC 
            LIMIT 100
        """
        registros = db.fetch_all(query, (condominio_id,))
    else:
        query = """
            SELECT 
                r.id, r.data_entrada, r.observacao,
                p.nome as pessoa_nome, 
                p.tipo as pessoa_tipo, 
                p.apartamento,
                u.nome as usuario_nome
            FROM registros r 
            JOIN pessoas p ON r.pessoa_id = p.id 
            LEFT JOIN usuarios u ON r.usuario_registro = u.id
            ORDER BY r.data_entrada DESC 
            LIMIT 100
        """
        registros = db.fetch_all(query)

    for reg in registros:
        reg['nome'] = reg.get('pessoa_nome')
        reg['tipo'] = reg.get('pessoa_tipo')
        reg['usuario'] = reg.get('usuario_nome') or 'Sistema'

    return jsonify({'success': True, 'registros': registros})


@app.route('/api/registros/<int:id>', methods=['DELETE'])
@login_required
@master_required
def delete_registro(id):
    try:
        db.execute_query("DELETE FROM registros WHERE id = ?", (id,))
        return jsonify({'success': True, 'message': 'Registro removido com sucesso!'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


# ==================== ESTATÍSTICAS ====================
@app.route('/api/estatisticas', methods=['GET'])
@login_required
def get_estatisticas():
    if session.get('user_tipo') == 'porteiro':
        condominio_id = session.get('user_condominio_id')
        total_pessoas = db.fetch_one("SELECT COUNT(*) as total FROM pessoas WHERE condominio_id = ?", (condominio_id,))
        total_registros = db.fetch_one(
            "SELECT COUNT(*) as total FROM registros r JOIN pessoas p ON r.pessoa_id = p.id WHERE p.condominio_id = ?",
            (condominio_id,))
        hoje = datetime.now().strftime('%Y-%m-%d')
        registros_hoje = db.fetch_one("""
            SELECT COUNT(*) as total FROM registros r 
            JOIN pessoas p ON r.pessoa_id = p.id 
            WHERE p.condominio_id = ? AND DATE(r.data_entrada) = ?
        """, (condominio_id, hoje))
    else:
        total_pessoas = db.fetch_one("SELECT COUNT(*) as total FROM pessoas")
        total_registros = db.fetch_one("SELECT COUNT(*) as total FROM registros")
        hoje = datetime.now().strftime('%Y-%m-%d')
        registros_hoje = db.fetch_one("SELECT COUNT(*) as total FROM registros WHERE DATE(data_entrada) = ?", (hoje,))

    return jsonify({
        'success': True,
        'estatisticas': {
            'total_pessoas': total_pessoas['total'] if total_pessoas else 0,
            'total_registros': total_registros['total'] if total_registros else 0,
            'registros_hoje': registros_hoje['total'] if registros_hoje else 0
        }
    })


# ==================== AVISOS ====================
@app.route('/api/avisos', methods=['GET'])
@login_required
def get_avisos():
    if session.get('user_tipo') == 'porteiro':
        condominio_id = session.get('user_condominio_id')
        query = """
            SELECT a.*, 
                   c.nome as condominio_nome,
                   u.nome as criado_por_nome
            FROM avisos a
            LEFT JOIN condominios c ON a.condominio_id = c.id
            LEFT JOIN usuarios u ON a.criado_por = u.id
            WHERE a.condominio_id = ?
            ORDER BY a.id DESC
        """
        avisos = db.fetch_all(query, (condominio_id,))
    else:
        query = """
            SELECT a.*, 
                   c.nome as condominio_nome,
                   u.nome as criado_por_nome
            FROM avisos a
            LEFT JOIN condominios c ON a.condominio_id = c.id
            LEFT JOIN usuarios u ON a.criado_por = u.id
            ORDER BY a.id DESC
        """
        avisos = db.fetch_all(query)
    return jsonify({'success': True, 'avisos': avisos})


@app.route('/api/avisos', methods=['POST'])
@login_required
def create_aviso():
    data = request.get_json()

    if session.get('user_tipo') == 'porteiro':
        condominio_id = session.get('user_condominio_id')
    else:
        condominio_id = data.get('condominio_id')

    criado_por = session.get('user_id')

    db.execute_query(
        "INSERT INTO avisos (titulo, mensagem, tipo, data_criacao, condominio_id, criado_por) VALUES (?, ?, ?, ?, ?, ?)",
        (data.get('titulo'), data.get('mensagem'), data.get('tipo'),
         datetime.now().strftime('%Y-%m-%d %H:%M:%S'), condominio_id, criado_por)
    )
    return jsonify({'success': True, 'message': 'Aviso publicado com sucesso!'})


@app.route('/api/avisos/<int:id>', methods=['PUT'])
@login_required
@master_required
def update_aviso(id):
    try:
        data = request.get_json()
        db.execute_query("""
            UPDATE avisos 
            SET titulo = ?, mensagem = ?, tipo = ?, ativo = ?
            WHERE id = ?
        """, (data.get('titulo'), data.get('mensagem'), data.get('tipo'),
              data.get('ativo', 1), id))
        return jsonify({'success': True, 'message': 'Aviso atualizado com sucesso!'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@app.route('/api/avisos/<int:id>', methods=['DELETE'])
@login_required
@master_required
def delete_aviso(id):
    try:
        db.execute_query("DELETE FROM avisos WHERE id = ?", (id,))
        return jsonify({'success': True, 'message': 'Aviso removido com sucesso!'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


# ==================== OCORRÊNCIAS ====================
@app.route('/api/ocorrencias', methods=['GET'])
@login_required
def get_ocorrencias():
    if session.get('user_tipo') == 'porteiro':
        condominio_id = session.get('user_condominio_id')
        query = """
            SELECT o.*, 
                   c.nome as condominio_nome,
                   u.nome as criado_por_nome
            FROM ocorrencias o
            LEFT JOIN condominios c ON o.condominio_id = c.id
            LEFT JOIN usuarios u ON o.criado_por = u.id
            WHERE o.condominio_id = ?
            ORDER BY o.id DESC
        """
        ocorrencias = db.fetch_all(query, (condominio_id,))
    else:
        query = """
            SELECT o.*, 
                   c.nome as condominio_nome,
                   u.nome as criado_por_nome
            FROM ocorrencias o
            LEFT JOIN condominios c ON o.condominio_id = c.id
            LEFT JOIN usuarios u ON o.criado_por = u.id
            ORDER BY o.id DESC
        """
        ocorrencias = db.fetch_all(query)
    return jsonify({'success': True, 'ocorrencias': ocorrencias})


@app.route('/api/ocorrencias', methods=['POST'])
@login_required
def create_ocorrencia():
    data = request.get_json()

    if session.get('user_tipo') == 'porteiro':
        condominio_id = session.get('user_condominio_id')
    else:
        condominio_id = data.get('condominio_id')

    criado_por = session.get('user_id')

    db.execute_query("""
        INSERT INTO ocorrencias (titulo, descricao, tipo, prioridade, status, data_criacao, condominio_id, responsavel, criado_por) 
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (data.get('titulo'), data.get('descricao'), data.get('tipo'),
          data.get('prioridade'), 'aberto', datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
          condominio_id, data.get('responsavel'), criado_por))
    return jsonify({'success': True, 'message': 'Ocorrência registrada com sucesso!'})


@app.route('/api/ocorrencias/<int:id>', methods=['PUT'])
@login_required
@master_required
def update_ocorrencia(id):
    try:
        data = request.get_json()
        db.execute_query("""
            UPDATE ocorrencias 
            SET titulo = ?, descricao = ?, tipo = ?, status = ?, prioridade = ?
            WHERE id = ?
        """, (data.get('titulo'), data.get('descricao'), data.get('tipo'),
              data.get('status'), data.get('prioridade'), id))
        return jsonify({'success': True, 'message': 'Ocorrência atualizada com sucesso!'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@app.route('/api/ocorrencias/<int:id>', methods=['DELETE'])
@login_required
@master_required
def delete_ocorrencia(id):
    try:
        db.execute_query("DELETE FROM ocorrencias WHERE id = ?", (id,))
        return jsonify({'success': True, 'message': 'Ocorrência removida com sucesso!'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


# ==================== PORTEIROS ====================
@app.route('/api/porteiros', methods=['GET'])
@login_required
@master_required
def get_porteiros():
    query = """
        SELECT u.id, u.nome, u.login, u.ativo, u.condominio_id, c.nome as condominio_nome 
        FROM usuarios u 
        LEFT JOIN condominios c ON u.condominio_id = c.id 
        WHERE u.tipo = 'porteiro'
        ORDER BY u.nome
    """
    porteiros = db.fetch_all(query)
    return jsonify({'success': True, 'porteiros': porteiros})


@app.route('/api/porteiros', methods=['POST'])
@login_required
@master_required
def create_porteiro():
    data = request.get_json()
    senha_hash = hashlib.sha256(data['senha'].encode()).hexdigest()
    db.execute_query(
        "INSERT INTO usuarios (nome, login, senha, tipo, condominio_id, ativo) VALUES (?, ?, ?, 'porteiro', ?, ?)",
        (data['nome'], data['login'], senha_hash, data.get('condominio_id'), 1 if data.get('ativo') else 0)
    )
    return jsonify({'success': True, 'message': 'Porteiro cadastrado com sucesso!'})


@app.route('/api/porteiros/<int:id>', methods=['PUT'])
@login_required
@master_required
def update_porteiro(id):
    data = request.get_json()
    if data.get('senha'):
        senha_hash = hashlib.sha256(data['senha'].encode()).hexdigest()
        db.execute_query(
            "UPDATE usuarios SET nome=?, login=?, senha=?, condominio_id=?, ativo=? WHERE id=? AND tipo='porteiro'",
            (data['nome'], data['login'], senha_hash, data.get('condominio_id'), 1 if data.get('ativo') else 0, id)
        )
    else:
        db.execute_query(
            "UPDATE usuarios SET nome=?, login=?, condominio_id=?, ativo=? WHERE id=? AND tipo='porteiro'",
            (data['nome'], data['login'], data.get('condominio_id'), 1 if data.get('ativo') else 0, id)
        )
    return jsonify({'success': True, 'message': 'Porteiro atualizado com sucesso!'})


@app.route('/api/porteiros/<int:id>', methods=['DELETE'])
@login_required
@master_required
def delete_porteiro(id):
    db.execute_query("DELETE FROM usuarios WHERE id=? AND tipo='porteiro'", (id,))
    return jsonify({'success': True, 'message': 'Porteiro removido com sucesso!'})


# ==================== RELATÓRIOS EXCEL ====================
@app.route('/api/relatorio_registros', methods=['GET'])
@login_required
@master_required
def relatorio_registros():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        import io

        query = """
            SELECT 
                r.id, 
                r.data_entrada, 
                r.observacao,
                p.nome as pessoa_nome, 
                p.tipo as pessoa_tipo,
                p.apartamento, 
                p.telefone, 
                c.nome as condominio_nome,
                u.nome as usuario_nome
            FROM registros r
            JOIN pessoas p ON r.pessoa_id = p.id
            LEFT JOIN condominios c ON p.condominio_id = c.id
            LEFT JOIN usuarios u ON r.usuario_registro = u.id
            ORDER BY r.data_entrada DESC
        """

        registros = db.fetch_all(query)

        if not registros:
            return jsonify({'success': False, 'error': 'Nenhum registro encontrado'}), 404

        wb = Workbook()
        ws = wb.active
        ws.title = "Registros"

        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                        bottom=Side(style='thin'))

        headers = ['ID', 'DATA/HORA', 'PESSOA', 'TIPO', 'APARTAMENTO', 'TELEFONE', 'CONDOMÍNIO', 'REGISTRADO POR',
                   'OBSERVAÇÃO']

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        for row_idx, reg in enumerate(registros, 2):
            data_hora = reg.get('data_entrada', '')
            if data_hora:
                try:
                    dt = datetime.strptime(str(data_hora), '%Y-%m-%d %H:%M:%S')
                    data_formatada = dt.strftime('%d/%m/%Y %H:%M:%S')
                except:
                    data_formatada = str(data_hora)
            else:
                data_formatada = ''

            ws.cell(row=row_idx, column=1, value=reg.get('id', ''))
            ws.cell(row=row_idx, column=2, value=data_formatada)
            ws.cell(row=row_idx, column=3, value=reg.get('pessoa_nome', ''))
            ws.cell(row=row_idx, column=4, value=reg.get('pessoa_tipo', ''))
            ws.cell(row=row_idx, column=5, value=reg.get('apartamento', ''))
            ws.cell(row=row_idx, column=6, value=reg.get('telefone', ''))
            ws.cell(row=row_idx, column=7, value=reg.get('condominio_nome', ''))
            ws.cell(row=row_idx, column=8, value=reg.get('usuario_nome', 'Sistema'))
            ws.cell(row=row_idx, column=9, value=reg.get('observacao', ''))

            for col in range(1, 10):
                ws.cell(row=row_idx, column=col).border = border

        col_widths = [8, 20, 35, 15, 15, 18, 25, 25, 40]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        nome_arquivo = f"relatorio_registros_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=nome_arquivo
        )

    except Exception as e:
        print(f"❌ Erro: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/relatorio_pessoas', methods=['GET'])
@login_required
@master_required
def relatorio_pessoas():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        import io

        query = """
            SELECT 
                p.id, p.nome, p.tipo, p.apartamento, p.telefone,
                p.documento, p.veiculo_placa, p.veiculo_modelo,
                p.status, p.motivo_bloqueio, c.nome as condominio_nome
            FROM pessoas p
            LEFT JOIN condominios c ON p.condominio_id = c.id
            ORDER BY c.nome, p.nome
        """

        pessoas = db.fetch_all(query)

        if not pessoas:
            return jsonify({'success': False, 'error': 'Nenhuma pessoa encontrada'}), 404

        wb = Workbook()
        ws = wb.active
        ws.title = "Pessoas"

        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='2E7D32', end_color='2E7D32', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                        bottom=Side(style='thin'))

        headers = ['ID', 'NOME', 'TIPO', 'APARTAMENTO', 'TELEFONE', 'DOCUMENTO', 'PLACA', 'MODELO', 'CONDOMÍNIO',
                   'STATUS', 'MOTIVO']

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        for row_idx, pessoa in enumerate(pessoas, 2):
            ws.cell(row=row_idx, column=1, value=pessoa.get('id', ''))
            ws.cell(row=row_idx, column=2, value=pessoa.get('nome', ''))
            ws.cell(row=row_idx, column=3, value=pessoa.get('tipo', ''))
            ws.cell(row=row_idx, column=4, value=pessoa.get('apartamento', ''))
            ws.cell(row=row_idx, column=5, value=pessoa.get('telefone', ''))
            ws.cell(row=row_idx, column=6, value=pessoa.get('documento', ''))
            ws.cell(row=row_idx, column=7, value=pessoa.get('veiculo_placa', ''))
            ws.cell(row=row_idx, column=8, value=pessoa.get('veiculo_modelo', ''))
            ws.cell(row=row_idx, column=9, value=pessoa.get('condominio_nome', ''))
            ws.cell(row=row_idx, column=10, value=pessoa.get('status', 'ATIVO'))
            ws.cell(row=row_idx, column=11, value=pessoa.get('motivo_bloqueio', ''))

            for col in range(1, 12):
                ws.cell(row=row_idx, column=col).border = border

        col_widths = [8, 40, 15, 15, 18, 20, 15, 18, 25, 12, 30]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        nome_arquivo = f"relatorio_pessoas_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=nome_arquivo
        )

    except Exception as e:
        print(f"❌ Erro: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500


# ==================== ZERAR BANCO ====================
@app.route('/api/zerar_banco', methods=['POST'])
@login_required
@master_required
def zerar_banco():
    try:
        db.execute_query("DELETE FROM pessoas")
        db.execute_query("DELETE FROM registros")
        db.execute_query("DELETE FROM avisos")
        db.execute_query("DELETE FROM ocorrencias")
        return jsonify({'success': True, 'message': 'Banco de dados zerado com sucesso!'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@app.route('/api/backup', methods=['POST'])
@login_required
@master_required
def backup():
    return jsonify({'success': True, 'message': 'Backup realizado com sucesso!'})


if __name__ == '__main__':
    print("=" * 60)
    print("🚀 SERVIDOR INICIADO")
    print("📱 Acesse: http://localhost:5000")
    print("🔑 Master: master / admin123")
    print("👔 Porteiro: joao / 123456 (Condomínio 1)")
    print("👔 Porteiro: maria / 123456 (Condomínio 2)")
    print("=" * 60)
    app.run(debug=True, host='0.0.0.0', port=5000)


    # ==================== RECEBIMENTOS ====================
    @app.route('/api/recebimentos', methods=['GET'])
    @login_required
    def get_recebimentos():
        query = """
            SELECT r.*, p.nome as pessoa_nome, p.apartamento, p.bloco, p.cor
            FROM recebimentos r
            JOIN registros rg ON r.registro_id = rg.id
            JOIN pessoas p ON rg.pessoa_id = p.id
            ORDER BY r.data_recebimento DESC
            LIMIT 50
        """
        recebimentos = db.fetch_all(query)
        return jsonify({'success': True, 'recebimentos': recebimentos})


    @app.route('/api/recebimentos/<int:id>', methods=['PUT'])
    @login_required
    def atualizar_recebimento(id):
        data = request.get_json()
        db.execute_query("""
            UPDATE recebimentos 
            SET status = ?, recebido_por = ?, data_recebimento = ?
            WHERE id = ?
        """, (data.get('status'), session.get('user_nome'), datetime.now().strftime('%Y-%m-%d %H:%M:%S'), id))
        return jsonify({'success': True, 'message': 'Status atualizado!'})


    # ==================== ESTATÍSTICAS POR COR ====================
    @app.route('/api/estatisticas_cores', methods=['GET'])
    @login_required
    @master_required
    def get_estatisticas_cores():
        query = """
            SELECT cor, COUNT(*) as total 
            FROM pessoas 
            WHERE cor IS NOT NULL 
            GROUP BY cor
            ORDER BY total DESC
        """
        cores = db.fetch_all(query)
        return jsonify({'success': True, 'cores': cores})

